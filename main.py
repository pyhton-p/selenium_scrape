import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, NoSuchWindowException
from openpyxl import Workbook
from flask import Flask, jsonify, request, send_file
from flask_mail import Mail, Message
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.styles import Alignment
app = Flask(__name__)
import getpass
import re

# Flask-Mail configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'Sender_mail'
app.config['MAIL_PASSWORD'] = 'p--- d--- i--- h ---'
app.config['MAIL_DEFAULT_SENDER'] = 'Sender_mail'

mail = Mail(app)
# Configure Firefox options
options = webdriver.FirefoxOptions()
# options.add_argument("--headless")

# # Start the Firefox WebDriver with the specified geckodriver path
driver = webdriver.Firefox(options=options)
driver.execute_script("window.scrollBy(0, 500);")

# # Set up WebDriverWait
wait = WebDriverWait(driver, 30)
options = webdriver.FirefoxOptions()
 
# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active


# Set column headers
ws.append(["Content Links", "Day","Month","Year", "email","Phone Number", "City", "Country", "Organization", "Job Title", "First Name","Last Name"])

# Function to periodically save data to Excel

def save_data():
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if '\n' in str(cell.value):
                cell.alignment = Alignment(wrap_text=True)
    wb.save("member_details.xlsx")
    print("Data saved successfully!")

# Function to send email with Excel file attached
def send_email_with_attachment(email, filename):
    from_email = "Sender_mail"
    password = "pdxi dukq nuih izsl"  # Your Gmail app password
    to_email = email

    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = 'Scraped Member Details'

    body = 'Please find the attached Excel file with scraped member details.'
    msg.attach(MIMEText(body, 'plain'))

    with open(filename, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )
    msg.attach(part)

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(from_email, password)
    text = msg.as_string()
    server.sendmail(from_email, to_email, text)
    server.quit()
       
    driver.quit()

def extract_date_components(input_string):
    # Define the specific pattern for "22 Dec 2024"
    pattern = r"(\d{1,2}) (\w{3}) (\d{4})"  # Matches "22 Dec 2024"
    cleaned_input = re.sub(r",.*$", "", input_string).strip()
    match = re.search(pattern, cleaned_input)
    if match:
        day, month, year = match.groups()
        return day.zfill(2), month, year  # Ensure day is two digits
    
    return None, None, None  # Return None if no patterns match


def scrape_member_info():
    members_with_content = []  # List to store member details with content links
    try:
        
        # Wait for the member list to be fully loaded
        wait.until(EC.presence_of_all_elements_located((By.XPATH, "//span[@class='member-list-text']")))
        
        # Fetch member elements
        member_elements = driver.find_elements(By.XPATH, "//span[@class='member-list-text']")
        
        # Iterate through each member
        for i in range(len(member_elements)):
            # Fetch member elements again to avoid StaleElementReferenceException
            member_elements = driver.find_elements(By.XPATH, "//span[@class='member-list-text']")
            member_element = member_elements[i]
            
            # Extract the full name and email address
            member_text = member_element.text.split('(')
            member_name = member_text[0].strip()
            email = member_text[1].split(')')[0].strip() if len(member_text) > 1 else ''
            
            name_parts = member_name.split()
            first_name = name_parts[0] if name_parts else ''
            last_name = name_parts[-1] if len(name_parts) > 1 else ''

            text_node_script = "return arguments[0].nextSibling.textContent.trim();"
            
            phone_number_element = member_element.find_element(By.XPATH, ".//strong[contains(text(), 'Phone Number:')]")
            phone_number = driver.execute_script(text_node_script, phone_number_element)
            
            city_element = member_element.find_element(By.XPATH, ".//strong[contains(text(), 'City:')]")
            city = driver.execute_script(text_node_script, city_element)
            
            country_element = member_element.find_element(By.XPATH, ".//strong[contains(text(), 'Country or US State:')]")
            country = driver.execute_script(text_node_script, country_element)
            
            organization_element = member_element.find_element(By.XPATH, ".//strong[contains(text(), 'Organization:')]")
            organization = driver.execute_script(text_node_script, organization_element)
            
            job_title_element = member_element.find_element(By.XPATH, ".//strong[contains(text(), 'Job Title:')]")
            job_title = driver.execute_script(text_node_script, job_title_element)

            # Extract member details page URL
            member_details_link = member_element.find_element(By.CLASS_NAME, "member-details-link").get_attribute("href")
            
            # Visit member details page
            driver.get(member_details_link)
            
            try:
                content_links_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//ms-button[text()='Content Links']")))
                content_links_button.click()
                time.sleep(2)
            except TimeoutException:
                print(f"Content Links button not found for {member_name}")
                continue  # Skip this member if content links button not found
            except Exception as e:
                print(f"An unexpected error occurred while clicking Content Links button for {member_name}: {e}")
                continue  # Skip this member if any unexpected error occurs
            
            try:
                content_links_elements = WebDriverWait(driver, 5).until(
                    EC.visibility_of_all_elements_located((By.XPATH, "//div[contains(@class, 'MemberSpaceWidgetInternal__MemberEventRow__3YZuZ__content')]"))
                )
                content_links = [element.text.strip() for element in content_links_elements if element.text.strip()]

                parent_container = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "MemberSpaceWidgetInternal__Pagination__07UsP__listContainer"))
                )
                rows = parent_container.find_elements(By.CLASS_NAME, "MemberSpaceWidgetInternal__MemberEventRow__3YZuZ__listRow")
                content_dates = []
                for row in rows:
                    try:

                        tooltip_trigger = WebDriverWait(row, 10).until(
                            EC.element_to_be_clickable((By.CLASS_NAME, "MemberSpaceWidgetInternal__MemberEventRow__3YZuZ__tooltipTrigger"))
                        )
                        # Scroll into view and click
                        #driver.execute_script("arguments[0].style.visibility = 'visible';", tooltip_trigger)
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tooltip_trigger)
                        time.sleep(1)
                        driver.execute_script("arguments[0].style.zIndex = '9999';", tooltip_trigger)
                        driver.execute_script("arguments[0].click();", tooltip_trigger)

                                                 
 
                        tooltip_trigger.click()
                        time.sleep(2)
                        
                        date_element = WebDriverWait(driver, 10,poll_frequency=1.0).until(
                            EC.presence_of_element_located((By.XPATH, "//div[span[text()='Date']]"))
                        )
                        date_text = date_element.text.split(":")[1].strip()
   
                        # date_element = driver.find_element(By.XPATH, "//div[span[text()='Date']]")
                        # date_text = date_element.text.split(":")[1].strip()
                        print(date_text)
                        content_dates.append(date_text)
                        time.sleep(1)
                    except Exception:
                        content_dates.append('')
                                 
           

                # Pair content links and dates: dalehoug@amazon.com
                processed_dates = [extract_date_components(date) for date in content_dates]               
                time.sleep(1)
                # Check if content links are available
                if content_links_elements:
                    content_links = [element.text.strip() for element in content_links_elements if element.text.strip() != '']

                    if content_links:
                        content_details = list(zip(content_links, processed_dates))
                        #print(content_details) 
                        # Create a separate row for each content link
                        for link, date_info in content_details:
                            date, month, year = date_info
                            ws.append([link, date, month, year, email, phone_number, city, country, organization, job_title, first_name, last_name])
                            print(f"Content Link: {link}, Day: {date}, Month: {month}, Year: {year},Email: {email}, Phone Number: {phone_number}, City: {city}, Country: {country}, Organization: {organization}, Job Title: {job_title}, First Name: {first_name}, Last Name: {last_name}")
                        print("-" * 50)
                        
                        # Save data periodically
                        if len(ws['A']) % 5 == 0:
                            time.sleep(1)
                            save_data()
                            time.sleep(1)
                             
                    else:
                        # If no content links, continue to next member
                        print(f"Skipping {first_name} {last_name} {email} - No content links.")
                        continue
                else:
                    print(f"No content links found for {first_name} {last_name} {email}.")  # Print message if no content links are found
                            
            except TimeoutException:
                print(f"{first_name} {last_name} {email} has no content links.")
                print("-" * 50)
            except Exception as e:
                print(f"An unexpected error occurred: {e}")
            # Go back to the main members page
            driver.back()
            time.sleep(1)
            # Wait for the member list to be fully loaded again
            wait.until(EC.presence_of_all_elements_located((By.XPATH, "//span[@class='member-list-text']")))

    except TimeoutException:
        print("Timeout occurred while waiting for member list to load")
    except Exception as e:
        print(f"An error occurred: {e}")
        # Save data when an error occurs
        save_data()
        time.sleep(1)
    
    return "member_details.xlsx"



def scrape_members():
    
    try:
        driver.get("https://admin.memberspace.com/sites/squarespace157/members?plan_status=free")
        if "sign_in" in driver.current_url:
            password = input("Enter your password: ")
            wait.until(EC.element_to_be_clickable((By.NAME, "email"))).send_keys("Login_mail")
            wait.until(EC.element_to_be_clickable((By.NAME, "password"))).send_keys(password)
            overlay = wait.until(EC.invisibility_of_element_located((By.ID, "__memberspace_modal_protected_page")))
            login_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//ms-button[text()='Log In']")))
            login_button.click()
        print("Reached the desired page after login")
        
        while True:
            all_member_info = []
            try:
                all_member_info.extend(scrape_member_info())
                next_page_button = driver.find_elements(By.XPATH, "//span[@class='next']/a")
                if not next_page_button:
                    break
                go_to_next_page()
                
            except (NoSuchWindowException, NoSuchElementException) as e:
                print("Next page button not found.")
                print("An error occurred while scraping member info:", e)
 
                break

                # Save data
        save_data()
        filename = "member_details.xlsx"
        
        # Send email with attachment
        try:
            send_email_with_attachment("Reciver_mail", filename)
        except Exception as e:
            print(f"An error occurred while sending email: {e}")

        return jsonify({"error": f"An error occurred: {str(e)}"}), 500
    
    except Exception as e:
        print(f"An error occurred while scraping member info: {e}")
        # Save data when an error occurs
        save_data()
      
        filename = "member_details.xlsx"
        
        # Send email with attachment
        try:
            send_email_with_attachment("Reciever_email", filename)
        except Exception as e:
            print(f"An error occurred while sending email: {e}")

        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

def go_to_next_page():
    try:
        next_page_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@class='next']/a")))
        next_page_button.click()
    except (NoSuchElementException, NoSuchWindowException) as e:
        print("An error occurred while navigating to the next page:", e)

if __name__ == "__main__":
    scrape_members()
